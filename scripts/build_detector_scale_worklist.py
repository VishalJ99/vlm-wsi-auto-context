#!/usr/bin/env python3
"""Build per-stain WSI worklists for detector-pipeline scale-up runs."""

from __future__ import annotations

import argparse
import csv
import json
import random
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_MANIFEST_CSV = Path("/data2/vj724/multistain/pilot_data/manifest.csv")
DEFAULT_SOURCE_WORKBOOK = Path(
    "/vol/biomedic3/histopatho/win_share/"
    "anon_master_combined_v5_with_reports_report_fixed (2).xlsx"
)
DEFAULT_STAINS = ("H&E", "PAS", "JONES", "EVG", "SV40")


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def repo_git_commit() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=REPO_ROOT,
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except Exception:
        return "unknown"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def read_workbook(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    forward_fill_keys = ("Anon_Patient_ID", "Anon_Path_ID", "biopsy_#", "biopsies_dates")
    forward_fill = {key: "" for key in forward_fill_keys}
    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            header: "" if index >= len(values) or values[index] is None else str(values[index]).strip()
            for index, header in enumerate(headers)
            if header
        }
        for key in forward_fill_keys:
            if row.get(key):
                forward_fill[key] = row[key]
            row[key] = forward_fill[key]
        row["case_id"] = row.get("Anon_Patient_ID", "")
        row["source_path"] = row.get("location_college", "")
        row["dest_path"] = ""
        row["slide_index_within_stain"] = row.get("slide_number", "")
        row["slide_index_within_case"] = row.get("slide_number", "")
        row["source_filename"] = Path(row.get("source_path", "")).name
        row["dest_filename"] = ""
        rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n")


def normalize_stain(value: str) -> str:
    stain = value.strip().upper()
    if stain in {"HNE", "H AND E", "H&E"}:
        return "H&E"
    return stain


def safe_stain(value: str) -> str:
    if value == "H&E":
        return "he"
    return re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_").lower()


def safe_token(value: str, fallback: str) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "", value.strip())
    return token or fallback


def numeric(value: str | None, default: int = 10**9) -> int:
    if value is None:
        return default
    value = str(value).strip()
    if not value:
        return default
    try:
        return int(float(value))
    except ValueError:
        return default


def existing_wsi_path(row: dict[str, str]) -> Path | None:
    for key in ("dest_path", "source_path"):
        value = row.get(key, "").strip()
        if not value:
            continue
        path = Path(value)
        if path.exists():
            return path
    return None


def row_sort_key(row: dict[str, str]) -> tuple[int, int, str]:
    return (
        numeric(row.get("slide_index_within_stain")),
        numeric(row.get("slide_index_within_case")),
        row.get("dest_path") or row.get("source_path") or "",
    )


def candidate_key(row: dict[str, str], stain: str) -> tuple[str, str, str]:
    return (
        row.get("case_id", "").strip(),
        row.get("Anon_Path_ID", "").strip(),
        stain,
    )


def representative_rows(rows: list[dict[str, str]], stains: tuple[str, ...]) -> dict[str, list[dict[str, str]]]:
    stain_set = {normalize_stain(stain) for stain in stains}
    grouped: dict[tuple[str, str, str], list[dict[str, str]]] = {}
    missing_path = 0
    for row in rows:
        stain = normalize_stain(row.get("Stain_from_wsi", ""))
        if stain not in stain_set:
            continue
        path = existing_wsi_path(row)
        if path is None:
            missing_path += 1
            continue
        key = candidate_key(row, stain)
        if not key[0] or not key[1]:
            continue
        copied = dict(row)
        copied["wsi_path"] = str(path)
        copied["normalized_stain"] = stain
        grouped.setdefault(key, []).append(copied)

    by_stain: dict[str, list[dict[str, str]]] = {stain: [] for stain in stain_set}
    for (_, _, stain), group_rows in grouped.items():
        by_stain[stain].append(sorted(group_rows, key=row_sort_key)[0])

    for stain in by_stain:
        by_stain[stain] = sorted(
            by_stain[stain],
            key=lambda row: (
                row.get("case_id", ""),
                row.get("Anon_Path_ID", ""),
                row_sort_key(row),
            ),
        )
    by_stain["_missing_path_rows"] = [{"count": str(missing_path)}]
    return by_stain


def materialize_link(row: dict[str, Any], link_root: Path, ordinal: int) -> Path:
    stain = str(row["stain"])
    stain_slug = safe_stain(stain)
    patient = safe_token(str(row.get("Anon_Patient_ID") or row.get("case_id") or ""), f"case{ordinal:04d}")
    path_id = safe_token(str(row.get("Anon_Path_ID") or ""), "path")
    slide = safe_token(str(row.get("slide_index_within_stain") or ""), f"{ordinal:04d}")
    source = Path(str(row["source_wsi_path"])).resolve()
    suffix = source.suffix if source.suffix else ".svs"
    filename = f"{stain_slug}_patient_{patient}{path_id}_slide_{slide}_{ordinal:04d}{suffix}"
    destination = link_root / stain / filename
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() or destination.is_symlink():
        destination.unlink()
    try:
        destination.symlink_to(source)
    except OSError:
        shutil.copy2(source, destination)
    return destination


def build_worklist(args: argparse.Namespace) -> int:
    manifest_csv = args.manifest_csv.resolve()
    source_workbook = args.source_workbook.resolve()
    output_dir = args.output_dir.resolve()
    if output_dir.exists() and not args.overwrite:
        raise FileExistsError(f"Output directory exists; pass --overwrite: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    stains = tuple(normalize_stain(stain) for stain in args.stains)
    if args.input_source == "workbook":
        rows = read_workbook(source_workbook)
        input_path = source_workbook
    else:
        rows = read_csv(manifest_csv)
        input_path = manifest_csv
    by_stain = representative_rows(rows, stains)
    missing_path_rows = int(by_stain.pop("_missing_path_rows")[0]["count"])
    rng = random.Random(args.seed)

    selected_rows: list[dict[str, Any]] = []
    counts: dict[str, dict[str, int]] = {}
    for stain in stains:
        candidates = by_stain.get(stain, [])
        if len(candidates) < args.per_stain:
            raise SystemExit(f"Need {args.per_stain} candidates for {stain}; found {len(candidates)}")
        sampled = list(candidates)
        rng.shuffle(sampled)
        selected = sorted(sampled[: args.per_stain], key=lambda row: row.get("wsi_path", ""))
        counts[stain] = {"eligible": len(candidates), "selected": len(selected)}
        for index, row in enumerate(selected, start=1):
            selected_rows.append(
                {
                    "selection_index_within_stain": index,
                    "stain": stain,
                    "case_id": row.get("case_id", ""),
                    "Anon_Patient_ID": row.get("Anon_Patient_ID", ""),
                    "Anon_Path_ID": row.get("Anon_Path_ID", ""),
                    "Anon_Slide_ID": row.get("Anon_Slide_ID", ""),
                    "slide_index_within_stain": row.get("slide_index_within_stain", ""),
                    "slide_index_within_case": row.get("slide_index_within_case", ""),
                    "source_wsi_path": row.get("wsi_path", ""),
                    "wsi_path": row.get("wsi_path", ""),
                    "dest_path": row.get("dest_path", ""),
                    "source_path": row.get("source_path", ""),
                    "dest_filename": row.get("dest_filename", ""),
                    "source_filename": row.get("source_filename", ""),
                    "manifest_csv": str(manifest_csv),
                    "source_workbook": str(source_workbook),
                    "selection_seed": args.seed,
                    "selection_rule": (
                        "Sample one representative WSI per case/path/stain, choosing the lowest "
                        "slide_index_within_stain then slide_index_within_case before deterministic "
                        "per-stain random sampling."
                    ),
                }
            )

    selected_rows = sorted(selected_rows, key=lambda row: (row["stain"], row["wsi_path"]))
    if args.link_root:
        link_root = args.link_root.resolve()
        for ordinal, row in enumerate(selected_rows, start=1):
            row["wsi_path"] = str(materialize_link(row, link_root, ordinal))
        link_root_value = str(link_root)
    else:
        link_root_value = ""

    for stain in stains:
        stain_rows = [row for row in selected_rows if row["stain"] == stain]
        list_path = output_dir / f"{safe_stain(stain)}_{args.per_stain}_wsi_list.txt"
        list_path.write_text("".join(f"{row['wsi_path']}\n" for row in stain_rows))

    fieldnames = list(selected_rows[0].keys()) if selected_rows else []
    write_csv(output_dir / "selected_wsis.csv", selected_rows, fieldnames)
    (output_dir / "all_500_wsi_list.txt").write_text(
        "".join(f"{row['wsi_path']}\n" for row in selected_rows)
    )
    non_sv40 = [row for row in selected_rows if row["stain"] != "SV40"]
    sv40 = [row for row in selected_rows if row["stain"] == "SV40"]
    (output_dir / "non_sv40_400_wsi_list.txt").write_text(
        "".join(f"{row['wsi_path']}\n" for row in non_sv40)
    )
    (output_dir / "sv40_100_wsi_list.txt").write_text(
        "".join(f"{row['wsi_path']}\n" for row in sv40)
    )

    summary = {
        "created_at": now_iso(),
        "repo_root": str(REPO_ROOT),
        "repo_commit": repo_git_commit(),
        "input_source": args.input_source,
        "input_path": str(input_path),
        "manifest_csv": str(manifest_csv),
        "source_workbook": str(source_workbook),
        "output_dir": str(output_dir),
        "link_root": link_root_value,
        "seed": args.seed,
        "per_stain": args.per_stain,
        "stains": list(stains),
        "missing_path_rows": missing_path_rows,
        "counts": counts,
        "outputs": {
            "selected_csv": str((output_dir / "selected_wsis.csv").resolve()),
            "all_wsi_list": str((output_dir / "all_500_wsi_list.txt").resolve()),
            "non_sv40_wsi_list": str((output_dir / "non_sv40_400_wsi_list.txt").resolve()),
            "sv40_wsi_list": str((output_dir / "sv40_100_wsi_list.txt").resolve()),
        },
    }
    write_json(output_dir / "summary.json", summary)
    command = " ".join([sys.executable, "scripts/build_detector_scale_worklist.py", *sys.argv[1:]])
    (output_dir / "reproduction.txt").write_text(
        f"""# Detector scale-up worklist reproduction

Generated: {summary['created_at']}
Repository: {REPO_ROOT}
Git commit: {summary['repo_commit']}

Command:
{command}

Manifest CSV: {manifest_csv}
Source workbook: {source_workbook}
Input source: {args.input_source} ({input_path})
Link root: {link_root_value or 'not used'}
Selection: {args.per_stain} WSIs per stain from {', '.join(stains)}
Seed: {args.seed}
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest-csv", type=Path, default=DEFAULT_MANIFEST_CSV)
    parser.add_argument("--source-workbook", type=Path, default=DEFAULT_SOURCE_WORKBOOK)
    parser.add_argument("--input-source", choices=["manifest", "workbook"], default="manifest")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--link-root", type=Path)
    parser.add_argument("--stains", nargs="+", default=list(DEFAULT_STAINS))
    parser.add_argument("--per-stain", type=int, default=100)
    parser.add_argument("--seed", type=int, default=244500)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    return build_worklist(args)


if __name__ == "__main__":
    raise SystemExit(main())
